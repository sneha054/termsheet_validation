
import pandas as pd
import openpyxl
import xlwings as xw

def extract_text_from_excel(path):
   data=pd.ExcelFile(path)#reading excel file
   print(data.sheet_names)
   all_txt=""
   for sheet in data.sheet_names:
    df=data.parse(sheet)
    print(f"----{sheet}----")
    print(df)
    all_txt="\n".join(df.astype(str).apply(" ".join,axis=1))
    print(all_txt)

def extract_text_from_excel_using_openpyxl(path):
   dataframe=openpyxl.load_workbook("Termsheet.xlsx")
   dataframe1=dataframe.active

   for row in range(0,dataframe1.max_row):
      for col in dataframe1.iter_cols(1,dataframe1.max_column):
         print(col[row].value)

def select_data_from_excel(path):
   data=pd.ExcelFile(path)
   for sheet in data.sheet_names:
      ws=xw.Book(path).sheets[sheet]
      v1=ws.range("A1:A7").value
      print("Result:",v1)
      
      
       
if __name__ == '__main__':
    excel_path='Termsheet.xlsx'
    extract_text_from_excel(excel_path)
    extract_text_from_excel_using_openpyxl(excel_path)
    select_data_from_excel(excel_path)